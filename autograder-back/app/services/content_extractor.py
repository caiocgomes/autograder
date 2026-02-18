"""
Content extraction from uploaded files for LLM evaluation.

Extracts text from PDF and XLSX files. Images are not extracted here;
they are passed directly as multimodal input to the LLM.
"""
from pathlib import Path

MAX_CONTENT_LENGTH = 50_000
TRUNCATION_NOTICE = "\n\n[Content truncated at 50,000 characters. Evaluation is based on partial content.]"


def extract_content(file_path: str | Path, content_type: str) -> str:
    """
    Extract text content from a file based on its content type.

    Args:
        file_path: Absolute path to the file
        content_type: MIME type of the file

    Returns:
        Extracted text content

    Raises:
        ValueError: For image types (should use multimodal input instead)
        ValueError: For unsupported content types
    """
    file_path = Path(file_path)

    if content_type == "application/pdf":
        text = _extract_pdf(file_path)
    elif content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        text = _extract_xlsx(file_path)
    elif content_type.startswith("image/"):
        raise ValueError(
            f"Image content type '{content_type}' should use multimodal LLM input, not text extraction"
        )
    else:
        raise ValueError(f"Unsupported content type for extraction: {content_type}")

    return _truncate(text)


def _extract_pdf(file_path: Path) -> str:
    """Extract text from PDF using pdfplumber, preserving table structure."""
    import pdfplumber

    pages = []
    with pdfplumber.open(file_path) as pdf:
        for i, page in enumerate(pdf.pages):
            # Extract tables first
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    rows = []
                    for row in table:
                        cells = [str(cell or "").strip() for cell in row]
                        rows.append(" | ".join(cells))
                    if rows:
                        # First row as header
                        header = rows[0]
                        separator = " | ".join(["---"] * len(table[0])) if table[0] else "---"
                        table_text = "\n".join([header, separator] + rows[1:])
                        pages.append(table_text)

            # Extract remaining text
            text = page.extract_text()
            if text:
                pages.append(text)

    return "\n\n".join(pages)


def _extract_xlsx(file_path: Path) -> str:
    """Extract data from XLSX, serializing each sheet as a markdown table."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Build markdown table
        lines = [f"### {sheet_name}"]

        # Header row
        header = rows[0]
        header_cells = [str(cell or "").strip() for cell in header]
        lines.append("| " + " | ".join(header_cells) + " |")
        lines.append("| " + " | ".join(["---"] * len(header_cells)) + " |")

        # Data rows
        for row in rows[1:]:
            cells = [str(cell or "").strip() for cell in row]
            lines.append("| " + " | ".join(cells) + " |")

        sheets.append("\n".join(lines))

    wb.close()
    return "\n\n".join(sheets)


def _truncate(text: str) -> str:
    """Truncate content to MAX_CONTENT_LENGTH with notice if truncated."""
    if len(text) <= MAX_CONTENT_LENGTH:
        return text
    return text[:MAX_CONTENT_LENGTH] + TRUNCATION_NOTICE
